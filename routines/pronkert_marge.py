#!/usr/bin/env python3
"""
PRONKERT MARGEFACTUUR — lezer.

Pronkert (onze backoffice/payroller) stuurt elke vrijdag een margefactuur per
mail: een creditfactuur waarop per flexkracht per dag staat wat wij aan die
uren verdienen. Dit bestand zet zo'n factuur om in nette regels, zodat de
routine ze in het CRM kan zetten zonder dat Tjeerd iets hoeft over te typen.

Twee bronnen, één uitkomst:
  • factuur-NNNNNN.pdf        — komt elke vrijdag, heeft een tekstlaag (geen OCR nodig)
  • Marge-overzicht ... .xlsx — komt af en toe van de relatiebeheerder, bevat
                                extra kolommen (klantnaam, klanttarief, factuurbedrag)

Beide leveren dezelfde regelvorm op, zodat de rest van de keten niet hoeft te
weten waar het vandaan kwam:

  {factuur, factuurdatum, week, jaar, weekmaandag, datum, naam, roepnaam,
   regnr, functie, klant, soort, uren, uurloon, tarief, factor, marge}

`marge` is POSITIEF = wat wij verdienen. Op de factuur staat het negatief
(creditbedrag dat Pronkert aan ons uitkeert); dat draaien we hier één keer om
zodat nergens anders in de code een minteken hoeft te staan.

Controle: de som van alle regels moet exact gelijk zijn aan het "Totaal" op de
factuur. Klopt dat niet, dan geeft lees_pdf() een waarschuwing terug in plaats
van stilletjes een verkeerd bedrag door te geven.
"""

import re
import datetime
from pathlib import Path

# ── Hulpjes ────────────────────────────────────────────────────────────────

def _getal(s):
    """'-1.077,26' → -1077.26 (Nederlands: punt = duizendtal, komma = decimaal)."""
    return float(str(s).replace('.', '').replace(',', '.'))


def _uren(s):
    """'8:00' → 8.0, '-8:00' → -8.0, '7:50' → 7.833… (minuten, geen honderdsten)."""
    neg = s.strip().startswith('-')
    u, m = s.strip().lstrip('-').split(':')
    w = int(u) + int(m) / 60
    return -w if neg else w


def maandag_van(jaar, week):
    """ISO-weeknummer → de maandag van die week (dat is de sleutel in fin_flex_weken)."""
    return datetime.date.fromisocalendar(int(jaar), int(week), 1).isoformat()


def splits_naam(ruw):
    """'S. van Nicolaas (Sven)' → ('S. van Nicolaas', 'Sven')."""
    m = re.match(r'\s*(.+?)\s*\(([^)]+)\)\s*$', str(ruw or ''))
    return (m.group(1).strip(), m.group(2).strip()) if m else (str(ruw or '').strip(), '')


# Alleen déze regels zijn gewerkte uren. Eindejaarsuitkering en
# arbeidstijdverkorting zijn reserveringen: ze leveren wél marge op, maar het
# zijn geen uren die meetellen voor de kosteloze-overnamegrens van de klant.
UURSOORTEN = ('loon normale uren', 'loon overwerkuren')


def is_gewerkt(soort):
    return str(soort or '').strip().lower() in UURSOORTEN


def persoon_sleutel(r):
    """Eén flexkracht = één sleutel, ongeacht de bron.

    NIET op registratienummer groeperen: dat staat wél op de PDF-factuur en
    níét op het Excel-marge-overzicht. Deed je dat toch, dan viel dezelfde
    persoon in tweeën — Sven kwam binnen als 32 uur (Excel) én als 114 uur
    (PDF), en de tweede helft overschreef de eerste. De naam staat op beide
    bronnen identiek: 'S. van Nicolaas (Sven)'.
    """
    ruw = f"{r.get('naam','')} {r.get('roepnaam','')}".lower()
    return '-'.join(re.findall(r'[a-z0-9]+', ruw))


# ── PDF ────────────────────────────────────────────────────────────────────

# Per pagina herhaalt de factuur een kop (adresblok) en een voet
# ("Transporteren pagina 2  -705,58"). Die knippen we eruit vóór het lezen,
# anders valt een regel die over een paginagrens loopt in tweeën.
_KOP = re.compile(r'Omschrijving\s+Uren/eenh\..*?ALPHEN AAN DEN RIJN', re.S)
_VOET = re.compile(r'(Factuurnummer\s+\d+\.\s*Transporteren pagina\s+\d+|Getransporteerd van pagina\s+\d+)\s+-?[\d.]*\d,\d{2}')

# Kop boven het blok regels van één flexkracht:
#   "Factuur 268110 S. van Nicolaas (Sven), Reg.nr. 7653911, Logistiek medewerker"
# Eén factuurregel:
#   "Week 31-2026 27-07-2026 Loon normale uren: 8:00, uurloon € 16,09 tarief
#    € 38,62 factor 1,7880   1,00   -78,80 2   -78,80"
_BLOK = re.compile(
    r'Factuur (?P<fnr>\d+) (?P<naam>[^,]+?\([^)]+\)), Reg\.nr\. (?P<reg>\d+), (?P<functie>.+?)(?= Week |$)'
    r'|Week (?P<wk>\d+)-(?P<jr>\d{4}) (?P<dat>\d{2}-\d{2}-\d{4}) (?P<soort>[^:]+?): (?P<uren>-?\d+:\d+),'
    r'(?P<rest>.*?) (?P<basis>-?[\d.]*\d,\d{2}) 2 (?P<bedrag>-?[\d.]*\d,\d{2})'
)


def lees_pdf(pad):
    import pypdf
    tekst = ' '.join((p.extract_text() or '') for p in pypdf.PdfReader(str(pad)).pages)
    return lees_tekst(tekst, Path(pad).name)


def lees_tekst(tekst, bestand=''):
    """Zelfde lezer, maar op kale tekst.

    De routine haalt de bijlage via Microsoft Graph op en krijgt de tekstlaag
    al uitgepakt terug. Dan hoeft er niets gedownload te worden — scheelt een
    bestand op de schijf en werkt ook als de mail alleen als tekst beschikbaar is.
    """
    tekst = re.sub(r'\s+', ' ', tekst or '')

    factuurnr = (re.search(r'Factuurnummer:\s*(\d+)', tekst) or [None, ''])[1]
    fdat = re.search(r'Factuurdatum:\s*(\d{2})-(\d{2})-(\d{4})', tekst)
    factuurdatum = f'{fdat.group(3)}-{fdat.group(2)}-{fdat.group(1)}' if fdat else ''
    tot = re.search(r'Totaal\s+(-?[\d.]*\d,\d{2})', tekst)
    totaal = -_getal(tot.group(1)) if tot else None   # omgedraaid: positief = onze marge

    schoon = _VOET.sub(' ', _KOP.sub(' ', tekst))
    schoon = re.sub(r'\s+', ' ', schoon)

    regels, huidig = [], None
    for m in _BLOK.finditer(schoon):
        if m.group('fnr'):
            naam, roep = splits_naam(m.group('naam'))
            huidig = dict(deelfactuur=m.group('fnr'), naam=naam, roepnaam=roep,
                          regnr=m.group('reg'), functie=m.group('functie').strip())
            continue
        if not huidig:
            continue                      # regel zonder kop: overslaan, niet gokken
        rest, soort = m.group('rest'), m.group('soort').strip()
        getal = lambda p: (_getal(re.search(p, rest).group(1)) if re.search(p, rest) else None)
        uren = _uren(m.group('uren'))
        uurloon, tarief = getal(r'uurloon € ([\d.,]+)'), getal(r'(?:basis)?tarief € ([\d.,]+)')
        # Omzet naar de klant staat niet apart op de factuur, maar volgt uit
        # tarief × uren. Daarmee kun je de marge ook als percentage van de
        # klantomzet zien, en niet alleen als bedrag.
        # Bij overwerk staat er een toeslagpercentage op de regel: de klant
        # betaalt dan basistarief × dat percentage. Laat je dat weg, dan valt de
        # klantomzet te laag uit (bij Alain in week 30 met €9,42).
        toeslag = getal(r'toeslag ([\d.,]+)%')
        klantbedrag = round(tarief * uren * ((toeslag / 100) if toeslag else 1), 2) \
            if (tarief is not None and uren) else None
        regels.append(dict(
            klantbedrag=klantbedrag,
            factuur=factuurnr, factuurdatum=factuurdatum, bron='pdf',
            week=int(m.group('wk')), jaar=int(m.group('jr')),
            weekmaandag=maandag_van(m.group('jr'), m.group('wk')),
            datum='{2}-{1}-{0}'.format(*m.group('dat').split('-')),
            soort=soort, uren=uren if is_gewerkt(soort) else 0.0, uren_regel=uren,
            uurloon=uurloon, tarief=tarief,
            factor=getal(r'factor ([\d.,]+)'),
            marge=-_getal(m.group('bedrag')),
            klant='',                      # staat niet op de PDF — komt uit het CRM
            **huidig))

    som = round(sum(r['marge'] for r in regels), 2)
    waarschuwing = None
    if totaal is not None and abs(som - totaal) > 0.005:
        waarschuwing = (f'Som van de regels (€{som:.2f}) wijkt af van het factuurtotaal '
                        f'(€{totaal:.2f}) — factuur {factuurnr} niet automatisch verwerken.')
    return dict(bestand=bestand, factuur=factuurnr, factuurdatum=factuurdatum,
                totaal=totaal, regels=regels, waarschuwing=waarschuwing)


# ── Excel (marge-overzicht) ────────────────────────────────────────────────

def lees_xlsx(pad):
    import openpyxl
    wb = openpyxl.load_workbook(str(pad), data_only=True)
    ws = wb[wb.sheetnames[0]]
    kop = [c.value for c in ws[1]]
    ix = {h: i for i, h in enumerate(kop)}
    nodig = ['Periode', 'Jaar', 'Flexwerker', 'Datum', 'Factuurtekst', 'Intermediair']
    ontbreekt = [k for k in nodig if k not in ix]
    if ontbreekt:
        return dict(bestand=Path(pad).name, regels=[], totaal=None,
                    waarschuwing='Kolommen ontbreken: ' + ', '.join(ontbreekt))

    def v(r, k):
        return r[ix[k]] if k in ix else None

    regels = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not v(r, 'Flexwerker'):
            continue
        naam, roep = splits_naam(v(r, 'Flexwerker'))
        soort = str(v(r, 'Factuurtekst') or '')
        minuten = v(r, 'Aantal minuten')
        uren = (minuten / 60) if isinstance(minuten, (int, float)) else 0.0
        dat = v(r, 'Datum')
        regels.append(dict(
            factuur=str(v(r, 'Factuurnummer') or ''), factuurdatum='', bron='xlsx',
            deelfactuur=str(v(r, 'Factuurnummer') or ''),
            week=int(v(r, 'Periode')), jaar=int(v(r, 'Jaar')),
            weekmaandag=maandag_van(v(r, 'Jaar'), v(r, 'Periode')),
            datum=dat.date().isoformat() if hasattr(dat, 'date') else str(dat or ''),
            naam=naam, roepnaam=roep, regnr='', functie=str(v(r, 'Functie') or ''),
            klant=str(v(r, 'Bedrijfsnaam') or ''),
            soort=soort, uren=uren if is_gewerkt(soort) else 0.0, uren_regel=uren,
            uurloon=v(r, 'Uurloon/belast'), tarief=v(r, 'Tarief'), factor=v(r, 'Factor'),
            klantbedrag=v(r, 'Factuurbedrag'),
            marge=float(v(r, 'Intermediair') or 0)))
    return dict(bestand=Path(pad).name, factuur='', factuurdatum='',
                totaal=round(sum(x['marge'] for x in regels), 2),
                regels=regels, waarschuwing=None)


def lees(pad):
    return lees_pdf(pad) if str(pad).lower().endswith('.pdf') else lees_xlsx(pad)


# ── Samenvatten ────────────────────────────────────────────────────────────

def per_week(regels):
    """Regels → per week één blok met totaal en een uitsplitsing per flexkracht.

    Een latere factuur kan een eerdere week corrigeren (dat gebeurde op factuur
    268245: eerst vijf dagen teruggeboekt, daarna opnieuw gefactureerd). Daarom
    tellen we altijd óp binnen week+persoon; de regels zelf zijn de waarheid.
    """
    weken = {}
    for r in regels:
        w = weken.setdefault(r['weekmaandag'], dict(
            week=r['weekmaandag'], weeknr=r['week'], jaar=r['jaar'],
            bedrag=0.0, uren=0.0, krachten={}))
        sleutel = persoon_sleutel(r)
        k = w['krachten'].setdefault(sleutel, dict(
            naam=r['naam'], roepnaam=r['roepnaam'], regnr=r['regnr'],
            functie=r['functie'], klant=r['klant'], uren=0.0, marge=0.0,
            klantomzet=0.0, _nu=0.0, _loon=0.0, _tar=0.0, _fac=0.0,
            uurloon=r.get('uurloon'), tarief=r.get('tarief'), factor=r.get('factor')))
        k['uren'] += r['uren']
        k['marge'] += r['marge']
        k['klantomzet'] += (r.get('klantbedrag') or 0)
        # Factoren wegen we over de normale uren: overwerk en de reserveringen
        # lopen op een eigen (lagere) factor en zouden het beeld vertroebelen.
        if str(r['soort']).strip().lower() == 'loon normale uren' and r['uren']:
            k['_nu'] += r['uren']
            k['_loon'] += (r.get('uurloon') or 0) * r['uren']
            k['_tar']  += (r.get('tarief') or 0) * r['uren']
            k['_fac']  += (r.get('factor') or 0) * r['uren']
        if not k['klant'] and r['klant']:
            k['klant'] = r['klant']
        w['uren'] += r['uren']
        w['bedrag'] += r['marge']
    for w in weken.values():
        w['bedrag'] = round(w['bedrag'], 2)
        w['uren'] = round(w['uren'], 2)
        w['flexkrachten'] = sum(1 for k in w['krachten'].values() if k['marge'] or k['uren'])
        for k in w['krachten'].values():
            k.update(factoren(k))
            k['uren'] = round(k['uren'], 2)
            k['marge'] = round(k['marge'], 2)
            k['klantomzet'] = round(k['klantomzet'], 2)
            k['marge_per_uur'] = round(k['marge'] / k['uren'], 2) if k['uren'] else None
            k['marge_pct'] = round(k['marge'] / k['klantomzet'], 4) if k['klantomzet'] else None
            for weg in ('_nu', '_loon', '_tar', '_fac'):
                k.pop(weg, None)
    return [weken[k] for k in sorted(weken)]


def factoren(k):
    """Wat de factuur over de opbouw van het uurtarief zegt.

    De factor op de regel is de INKOOPfactor van Pronkert: uurloon × factor is
    wat zij ons per uur rekenen. Het tarief op dezelfde regel is wat de klant
    per uur betaalt. Daaruit volgt de klantfactor, en het verschil tussen die
    twee is de marge — uitgedrukt in factor in plaats van in euro's.

        inkooptarief = uurloon × inkoopfactor
        klantfactor  = klanttarief ÷ uurloon
        margefactor  = klantfactor − inkoopfactor
        marge/uur    = margefactor × uurloon
    """
    nu = k.get('_nu') or 0
    if not nu:
        return dict(uurloon_gem=None, inkoopfactor=None, klantfactor=None,
                    margefactor=None, inkooptarief=None, klanttarief=None)
    loon = k['_loon'] / nu
    tar  = k['_tar'] / nu
    fac  = k['_fac'] / nu
    return dict(
        uurloon_gem=round(loon, 2),
        inkoopfactor=round(fac, 4) or None,
        klanttarief=round(tar, 2),
        inkooptarief=round(loon * fac, 2) if fac else None,
        klantfactor=round(tar / loon, 4) if loon else None,
        margefactor=round(tar / loon - fac, 4) if (loon and fac) else None)


def euro(n):
    return '€' + f'{n:,.2f}'.replace(',', '~').replace('.', ',').replace('~', '.')


if __name__ == '__main__':
    import sys, json

    args = sys.argv[1:]
    als_json = '--json' in args
    args = [a for a in args if a != '--json']

    alles, waarschuwingen, facturen = [], [], []
    i = 0
    while i < len(args):
        a = args[i]
        if a == '--tekst':                       # kale tekstlaag (uit Microsoft Graph)
            i += 1
            d = lees_tekst(Path(args[i]).read_text(), Path(args[i]).name)
        else:
            d = lees(a)
        if d.get('waarschuwing'):
            waarschuwingen.append(f"{d['bestand']}: {d['waarschuwing']}")
        if d.get('factuur'):
            facturen.append(d['factuur'])
        alles += d['regels']
        i += 1

    weken = per_week(alles)
    if als_json:
        print(json.dumps(dict(regels=alles, weken=weken, facturen=facturen,
                              waarschuwingen=waarschuwingen), ensure_ascii=False))
    else:
        for w in waarschuwingen:
            print('LET OP:', w)
        for w in weken:
            print(f"week {w['weeknr']} ({w['week']}) — {euro(w['bedrag'])} marge, "
                  f"{w['uren']:.2f} uur, {w['flexkrachten']} flexkrachten")
            for k in w['krachten'].values():
                pu = f" ({euro(k['marge_per_uur'])}/uur)" if k['marge_per_uur'] else ''
                print(f"   • {k['roepnaam'] or k['naam']}: {k['uren']:.2f} uur, "
                      f"{euro(k['marge'])}{pu}")
                if k['inkoopfactor']:
                    print(f"       uurloon {euro(k['uurloon_gem'])} · inkoop {k['inkoopfactor']:.4f} "
                          f"({euro(k['inkooptarief'])}) → klant {k['klantfactor']:.4f} "
                          f"({euro(k['klanttarief'])}) · marge-factor {k['margefactor']:.4f}"
                          + (f" · {k['marge_pct']*100:.1f}% van de klantomzet" if k['marge_pct'] else ''))
